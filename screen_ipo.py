#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Post IPO Monitor
================
2년 이내 신규상장 종목의 수급/거래량/RSI/변동성 모니터링

데이터 소스:
    - 최초상장일.xlsx: IPO 종목 리스트
    - 수급.xlsx: 기관/외국인 순매수 데이터
    - Bloomberg API: 가격, 거래량, RSI, 변동성

실행:
    python run.py
"""
import pandas as pd
from datetime import datetime
from pathlib import Path
import re

from post_ipo_daily import Config
from post_ipo_daily.utils import print_progress_bar, get_previous_business_day, get_today_business_day


class IPOMonitor:
    """IPO 종목 모니터링 클래스"""

    def __init__(self, config: Config = None):
        """
        Args:
            config: Config 인스턴스
        """
        self.config = config or Config()
        self.config.ensure_directories()

        # 데이터 저장
        self._ipo_df = None
        self._supply_df = None
        self._supply_date = None
        self._bloomberg_df = None
        self._result_df = None
        self._ref_date = None
        self._rsi_cache = {}  # {코드: RSI(14) 시계열} — RSI추이 히스토리 재사용용

    # =========================================================================
    # Step 1: IPO 종목 로드
    # =========================================================================
    def load_ipo_universe(self, source: str = 'A') -> pd.DataFrame:
        """IPO Universe 로드
        source='A': 최초상장일.xlsx 기준 (2년 이내 자동 필터)
        source='B': __post ipo univ.xlsx 기준 (Symbol 열 직접 사용)
        """
        print("\n" + "=" * 60)
        print("[Step 1] IPO Universe 로드")
        print("=" * 60)

        if source == 'B':
            return self._load_ipo_universe_b()
        return self._load_ipo_universe_a()

    def _load_ipo_universe_a(self) -> pd.DataFrame:
        """버전 A: 최초상장일.xlsx — 2년 이내 상장종목 자동 필터"""
        df = pd.read_excel(self.config.IPO_FILE, skiprows=5)
        df.columns = ['코드', '코드명', '최초상장일', '상장일']

        # datetime 변환
        df['최초상장일_dt'] = pd.to_datetime(
            df['최초상장일'].fillna(0).astype(int).astype(str),
            format='%Y%m%d',
            errors='coerce'
        )

        # 2년 이내 필터링
        from datetime import timedelta
        cutoff = datetime.now() - timedelta(days=self.config.IPO_DAYS_LIMIT)
        recent = df[df['최초상장일_dt'] >= cutoff].copy()
        recent = recent.dropna(subset=['코드'])

        # A+6자리 숫자만
        def is_regular_code(code):
            return bool(re.match(r'^A\d{6}$', str(code)))

        regular = recent[recent['코드'].apply(is_regular_code)].copy()

        # ETF/스팩/리츠 제외
        def is_regular_stock(name):
            if not isinstance(name, str):
                return True
            name_upper = name.upper()
            for kw in self.config.EXCLUDE_KEYWORDS:
                if kw.upper() in name_upper:
                    return False
            return True

        stocks = regular[regular['코드명'].apply(is_regular_stock)].copy()

        stocks = stocks[['코드', '코드명', '최초상장일', '최초상장일_dt']].copy()
        stocks = stocks.drop_duplicates(subset=['코드'], keep='first')
        stocks['days_since_ipo'] = (datetime.now() - stocks['최초상장일_dt']).dt.days
        stocks['ticker_ks'] = stocks['코드'].str[1:] + ' KS Equity'
        stocks['ticker_kq'] = stocks['코드'].str[1:] + ' KQ Equity'
        stocks = stocks.sort_values('최초상장일_dt', ascending=False)
        stocks = stocks.reset_index(drop=True)

        self._ipo_df = stocks
        print(f"[버전 A] 기준일: {cutoff.strftime('%Y-%m-%d')} 이후 상장")
        print(f"대상 종목: {len(stocks)}개")
        return stocks

    def _load_ipo_universe_b(self) -> pd.DataFrame:
        """버전 B: __post ipo univ.xlsx — Symbol 열의 종목 직접 사용"""
        df = pd.read_excel(self.config.UNIV_FILE, header=1)

        # Symbol, Name 컬럼 확인
        if 'Symbol' not in df.columns:
            print("오류: __post ipo univ.xlsx에 'Symbol' 컬럼이 없습니다.")
            return pd.DataFrame()

        df = df[['Symbol', 'Name']].copy()
        df.columns = ['코드', '코드명']
        df = df.dropna(subset=['코드'])

        # A+6자리 숫자만 (표준 종목코드)
        def is_regular_code(code):
            return bool(re.match(r'^A\d{6}$', str(code)))

        stocks = df[df['코드'].apply(is_regular_code)].copy()

        # 상장일 정보 없음 → NaT/NaN 처리
        stocks['최초상장일'] = None
        stocks['최초상장일_dt'] = pd.NaT
        stocks['days_since_ipo'] = None
        stocks['ticker_ks'] = stocks['코드'].str[1:] + ' KS Equity'
        stocks['ticker_kq'] = stocks['코드'].str[1:] + ' KQ Equity'
        stocks = stocks.reset_index(drop=True)

        self._ipo_df = stocks
        print(f"[버전 B] __post ipo univ.xlsx 로드")
        print(f"대상 종목: {len(stocks)}개")
        return stocks

    # =========================================================================
    # Step 2: 수급 데이터 로드
    # =========================================================================
    def load_supply_data(self) -> pd.DataFrame:
        """수급.xlsx에서 기관/외국인 순매수 데이터 로드"""
        print("\n" + "=" * 60)
        print("[Step 2] 수급 데이터 로드")
        print("=" * 60)

        # 전체 데이터 읽기
        raw = pd.read_excel(self.config.SUPPLY_FILE, header=None)

        # 메타데이터 추출 (행 8~13)
        codes = raw.iloc[8, :].values
        names = raw.iloc[9, :].values
        item_codes = raw.iloc[11, :].values

        # 데이터 부분 (행 14부터)
        data = raw.iloc[14:, :].copy()
        data.columns = range(len(data.columns))

        # 날짜 컬럼 (0번)
        data[0] = pd.to_datetime(data[0], errors='coerce')
        data = data.dropna(subset=[0])
        data = data.rename(columns={0: '날짜'})

        # 가장 최근 날짜 데이터만 추출
        latest_date = data['날짜'].max()
        latest = data[data['날짜'] == latest_date].iloc[0]

        self._supply_date = latest_date
        print(f"최신 데이터 날짜: {latest_date.strftime('%Y-%m-%d')}")

        # 종목별 수급 데이터 정리
        supply_records = []

        i = 1
        while i < len(codes):
            code = codes[i]
            if pd.isna(code) or not str(code).startswith('A'):
                i += 1
                continue

            record = {'코드': code, '코드명': names[i]}

            j = i
            while j < len(codes) and codes[j] == code:
                item_code = item_codes[j]
                value = latest[j] if j < len(latest) else None

                if item_code == 'CI20003020':
                    record['기관_일간'] = value
                elif item_code == 'CI20003021':
                    record['기관_5일'] = value
                elif item_code == 'CI20003022':
                    record['기관_20일'] = value
                elif item_code == 'CI20113020':
                    record['외국인_일간'] = value
                elif item_code == 'CI20113021':
                    record['외국인_5일'] = value
                elif item_code == 'CI20113022':
                    record['외국인_20일'] = value

                j += 1

            supply_records.append(record)
            i = j

        supply_df = pd.DataFrame(supply_records)

        # 중복 코드 제거 (수급.xlsx에 동일 종목 중복 시)
        supply_df = supply_df.drop_duplicates(subset=['코드'], keep='first')

        # 숫자 컬럼 변환
        numeric_cols = ['기관_일간', '기관_5일', '기관_20일', '외국인_일간', '외국인_5일', '외국인_20일']
        for col in numeric_cols:
            if col in supply_df.columns:
                supply_df[col] = pd.to_numeric(supply_df[col], errors='coerce')

        self._supply_df = supply_df

        print(f"수급 데이터: {len(supply_df)}개 종목")

        return supply_df

    # =========================================================================
    # Step 3: 가격/거래량 데이터 수집 (당일 장 마감 기준)
    #   - Bloomberg(blpapi) 제거 → FinanceDataReader(한국 증시 공개 데이터) 기반
    # =========================================================================
    @staticmethod
    def _wilder_rsi(close: pd.Series, length: int = 14) -> pd.Series:
        """Wilder 방식 RSI — Bloomberg RSI_14D와 동일한 평활(RMA) 사용"""
        delta = close.diff()
        gain = delta.clip(lower=0)
        loss = -delta.clip(upper=0)
        avg_gain = gain.ewm(alpha=1 / length, min_periods=length, adjust=False).mean()
        avg_loss = loss.ewm(alpha=1 / length, min_periods=length, adjust=False).mean()
        rs = avg_gain / avg_loss
        return (100 - 100 / (1 + rs)).round(2)

    @staticmethod
    def _load_market_caps() -> dict:
        """fdr.StockListing('KRX')로 종목별 시가총액(억원) 맵 생성"""
        import FinanceDataReader as fdr
        try:
            listing = fdr.StockListing('KRX')
        except Exception as e:
            print(f"  시가총액 로드 실패(StockListing): {e}")
            return {}
        code_col = next((c for c in ['Code', 'code', 'Symbol'] if c in listing.columns), None)
        cap_col = next((c for c in ['Marcap', 'MarCap', 'marcap', '시가총액'] if c in listing.columns), None)
        if code_col is None or cap_col is None:
            print(f"  시가총액 컬럼 탐색 실패 (컬럼: {list(listing.columns)[:10]})")
            return {}
        caps = {}
        for _, r in listing[[code_col, cap_col]].dropna().iterrows():
            try:
                caps[str(r[code_col]).zfill(6)] = float(r[cap_col]) / 100_000_000  # 원 → 억
            except (ValueError, TypeError):
                continue
        return caps

    def fetch_bloomberg_data(self, use_prev_day: bool = False) -> pd.DataFrame:
        """FinanceDataReader(한국 증시 전용)로 가격, 거래량, RSI, 변동성 수집
        use_prev_day=False: 당일(T-0) 종가 기준 (장 마감 직후 실행)
        use_prev_day=True:  전일(T-1) 종가 기준 (다음날 실행)
        """
        import FinanceDataReader as fdr

        mode_label = '전일 종가 기준 (다음날 실행)' if use_prev_day else '당일 장 마감 기준'
        print("\n" + "=" * 60)
        print(f"[Step 3] 가격/거래량 데이터 수집 ({mode_label})")
        print("=" * 60)

        if self._ipo_df is None:
            print("오류: IPO Universe를 먼저 로드하세요.")
            return pd.DataFrame()

        # 기준일 결정
        if use_prev_day:
            ref_date = get_previous_business_day(self.config)
        else:
            ref_date = get_today_business_day(self.config)
        ref_date_str = ref_date.strftime('%Y-%m-%d')
        self._ref_date = ref_date

        print(f"기준일: {ref_date_str} ({'전일 종가' if use_prev_day else '당일 종가'})")

        # 수급 데이터 날짜 검증
        if use_prev_day:
            # 다음날 실행: 수급 데이터도 ref_date(전일) 기준이어야 함
            expected_supply_str = ref_date_str
        else:
            # 당일 실행: 수급은 T-1 기준
            expected_supply_date = get_previous_business_day(self.config)
            expected_supply_str = expected_supply_date.strftime('%Y-%m-%d')

        if self._supply_date is not None:
            supply_date_str = self._supply_date.strftime('%Y-%m-%d')
            if supply_date_str != expected_supply_str:
                print(f"  ⚠️  수급 데이터 날짜({supply_date_str})가 기대값({expected_supply_str})과 불일치!")
                print(f"      수급.xlsx 파일을 확인하세요.")
            else:
                print(f"  수급 기준일: {supply_date_str}  |  주가 기준일: {ref_date_str}")

        # 종목코드 추출 (A000000 → 000000)
        stock_codes = self._ipo_df['코드'].str[1:].tolist()
        print(f"대상: {len(stock_codes)}개 종목")

        # 시가총액(억) 맵 — 1000억 필터용
        mkt_caps = self._load_market_caps()
        print(f"시가총액 로드: {len(mkt_caps)}개\n")

        # 기술적 지표 계산을 위해 90일(달력) 히스토리 확보
        start_date = (ref_date - pd.Timedelta(days=90)).strftime('%Y-%m-%d')
        end_date = ref_date_str

        results = []
        self._rsi_cache = {}
        batch_size = self.config.BATCH_SIZE
        total_batches = (len(stock_codes) + batch_size - 1) // batch_size

        for i in range(0, len(stock_codes), batch_size):
            batch_codes = stock_codes[i:i + batch_size]
            batch_num = i // batch_size + 1
            print_progress_bar(batch_num, total_batches, prefix='데이터 수집')

            for code in batch_codes:
                try:
                    # FinanceDataReader: KRX 접두형 우선, 실패 시 기본 코드로 재시도
                    hist = None
                    for symbol in (f'KRX:{code}', code):
                        try:
                            hist = fdr.DataReader(symbol, start_date, end_date)
                            if hist is not None and not hist.empty:
                                break
                        except Exception:
                            hist = None

                    if hist is None or hist.empty or len(hist) < 5:
                        continue

                    hist.columns = hist.columns.str.lower()
                    if 'close' not in hist.columns or 'volume' not in hist.columns:
                        continue

                    close = hist['close']
                    volume = hist['volume']

                    # 기술적 지표
                    hist['rsi_14d'] = self._wilder_rsi(close, 14)
                    hist['mov_avg_10d'] = close.rolling(10).mean()
                    hist['mov_avg_20d'] = close.rolling(20).mean()
                    hist['mov_avg_60d'] = close.rolling(60).mean()

                    # 볼린저밴드 폭 (20일, ±2σ) = 상단 - 하단 = 4σ
                    sd20 = close.rolling(20).std(ddof=0)
                    hist['bollinger_band_width'] = (sd20 * 4).round(2)

                    # 변동성: 일간수익률 30일 표준편차(%)
                    hist['volatility_30d'] = (close.pct_change().rolling(30).std() * 100).round(2)

                    # 거래량 지표
                    hist['volume_avg_20d'] = volume.rolling(20).mean()
                    hist['rvol_20'] = (volume / hist['volume_avg_20d']).round(2)

                    # 수익률
                    hist['chg_pct_1d'] = (close.pct_change() * 100).round(2)
                    hist['chg_pct_20d'] = (((close - close.shift(20)) / close.shift(20)) * 100).round(2)

                    # RSI추이 히스토리 재사용을 위해 RSI 시계열 캐시
                    self._rsi_cache[code] = hist['rsi_14d']

                    # 기준일 위치 결정 (당일=마지막, 전일=ref_date 이하 마지막)
                    if use_prev_day:
                        positions = [k for k, d in enumerate(hist.index)
                                     if d.strftime('%Y-%m-%d') <= ref_date_str]
                        if not positions:
                            continue
                        pos = positions[-1]
                    else:
                        pos = len(hist) - 1

                    row = hist.iloc[pos]
                    cur_rsi = row.get('rsi_14d')
                    prev_rsi = hist['rsi_14d'].iloc[pos - 1] if pos >= 1 else None
                    if pd.notna(cur_rsi) and prev_rsi is not None and pd.notna(prev_rsi):
                        rsi_change = round(float(cur_rsi) - float(prev_rsi), 1)
                    else:
                        rsi_change = None

                    close_v = row.get('close')
                    vol_v = row.get('volume')
                    turnover = (round((close_v * vol_v) / 100_000_000, 1)
                                if (pd.notna(close_v) and pd.notna(vol_v)) else None)

                    results.append({
                        'ticker': code,
                        'px_last': close_v,
                        'px_volume': vol_v,
                        'turnover': turnover,
                        'chg_pct_1d': row.get('chg_pct_1d'),
                        'chg_pct_20d': row.get('chg_pct_20d'),
                        'rsi_14d': cur_rsi,
                        'rsi_prev': prev_rsi,
                        'rsi_change': rsi_change,
                        'volume_avg_20d': row.get('volume_avg_20d'),
                        'rvol_20': row.get('rvol_20'),
                        'volatility_30d': row.get('volatility_30d'),
                        'mov_avg_10d': row.get('mov_avg_10d'),
                        'mov_avg_20d': row.get('mov_avg_20d'),
                        'mov_avg_60d': row.get('mov_avg_60d'),
                        'bollinger_band_width': row.get('bollinger_band_width'),
                        'cur_mkt_cap': mkt_caps.get(code),
                    })
                except Exception:
                    continue

        if not results:
            print("\n데이터 수집 실패")
            return pd.DataFrame()

        df = pd.DataFrame(results).set_index('ticker')

        # 이동평균선 돌파 여부
        if 'px_last' in df.columns and 'mov_avg_10d' in df.columns:
            df['above_ma10'] = (df['px_last'] > df['mov_avg_10d']).astype(int)
        if 'px_last' in df.columns and 'mov_avg_20d' in df.columns:
            df['above_ma20'] = (df['px_last'] > df['mov_avg_20d']).astype(int)
        if 'px_last' in df.columns and 'mov_avg_60d' in df.columns:
            df['above_ma60'] = (df['px_last'] > df['mov_avg_60d']).astype(int)

        self._bloomberg_df = df

        print(f"\n수집 완료: {len(df)}개 종목 (기준일: {ref_date_str})")

        # 원본 데이터 저장
        self._save_bloomberg_raw_data(ref_date_str)

        return self._bloomberg_df

    def _save_bloomberg_raw_data(self, ref_date_str: str):
        """Bloomberg 원본 데이터를 Excel로 저장"""
        if self._bloomberg_df is None or self._bloomberg_df.empty:
            return

        # raw data 폴더 생성
        raw_data_dir = self.config.OUTPUT_DIR / "raw data"
        raw_data_dir.mkdir(parents=True, exist_ok=True)

        filename = f"bloomberg_raw_{ref_date_str.replace('-', '')}.xlsx"
        filepath = raw_data_dir / filename

        try:
            self._bloomberg_df.reset_index().to_excel(filepath, index=False, sheet_name='RawData')
            print(f"  원본 데이터 저장: {filepath}")
        except Exception as e:
            print(f"  Bloomberg 원본 저장 실패: {e}")

    # =========================================================================
    # Step 4: 데이터 병합 및 최종 결과 생성
    # =========================================================================
    def merge_data(self) -> pd.DataFrame:
        """IPO + 수급 + Bloomberg 데이터 병합"""
        print("\n" + "=" * 60)
        print("[Step 4] 데이터 병합")
        print("=" * 60)

        if self._ipo_df is None:
            print("오류: IPO Universe가 없습니다.")
            return pd.DataFrame()

        # 기본 IPO 데이터
        result = self._ipo_df[['코드', '코드명', '최초상장일_dt', 'days_since_ipo', 'ticker_ks']].copy()
        result = result.rename(columns={'최초상장일_dt': '상장일'})

        # 수급 데이터 병합
        if self._supply_df is not None and not self._supply_df.empty:
            supply_cols = ['코드', '기관_일간', '기관_5일', '기관_20일', '외국인_일간', '외국인_5일', '외국인_20일']
            supply_cols = [c for c in supply_cols if c in self._supply_df.columns]
            result = result.merge(self._supply_df[supply_cols], on='코드', how='left')

        # Bloomberg 데이터 병합
        if self._bloomberg_df is not None and not self._bloomberg_df.empty:
            bbg = self._bloomberg_df.copy()
            bbg['코드'] = 'A' + bbg.index.astype(str)

            bbg_cols = ['코드', 'px_last', 'px_volume', 'turnover', 'volume_avg_20d', 'rvol_20',
                        'volatility_30d', 'rsi_14d', 'rsi_prev', 'rsi_change',
                        'chg_pct_1d', 'chg_pct_20d',
                        'mov_avg_10d', 'mov_avg_20d', 'mov_avg_60d',
                        'above_ma10', 'above_ma20', 'above_ma60',
                        'bollinger_band_width',
                        'cur_mkt_cap', 'eqy_free_float_pct', 'short_int_ratio']
            bbg_cols = [c for c in bbg_cols if c in bbg.columns]

            result = result.merge(bbg[bbg_cols], on='코드', how='left')

        # 중복 행 제거 (merge 후 혹시 남은 중복)
        result = result.drop_duplicates(subset=['코드'], keep='first')

        # 컬럼명 정리
        rename_map = {
            'px_last': '현재가',
            'px_volume': '거래량',
            'turnover': '거래대금(억)',
            'volume_avg_20d': '20일평균거래량',
            'rvol_20': 'RVOL(20)',
            'volatility_30d': '변동성(30D)',
            'bollinger_band_width': 'BB폭',
            'rsi_14d': 'RSI(14)',
            'rsi_prev': '전일RSI(14)',
            'rsi_change': 'RSI변화량',
            'chg_pct_1d': '1일수익률(%)',
            'chg_pct_20d': '20일수익률(%)',
            'mov_avg_10d': '10일이평',
            'mov_avg_20d': '20일이평',
            'mov_avg_60d': '60일이평',
            'above_ma10': '10일선돌파',
            'above_ma20': '20일선돌파',
            'above_ma60': '60일선돌파',
            'cur_mkt_cap': '시가총액(억)',
            'eqy_free_float_pct': '유통비율(%)',
            'short_int_ratio': '공매도비율',
        }
        result = result.rename(columns=rename_map)

        # 시가총액 1000억 미만 제외 (FinanceDataReader StockListing 기준, 억원)
        if '시가총액(억)' in result.columns:
            result['시가총액(억)'] = pd.to_numeric(result['시가총액(억)'], errors='coerce').round(0)
            before = len(result)
            result = result[result['시가총액(억)'].fillna(0) >= 1000].copy()
            print(f"  시가총액 1000억 미만 제외: {before - len(result)}개 종목 제거 → {len(result)}개 유지")

        # 상장일 형식 변경
        if '상장일' in result.columns:
            result['상장일'] = pd.to_datetime(result['상장일']).dt.strftime('%Y-%m-%d')

        # 단위 변환 (백만원)
        for col in ['기관_일간', '기관_5일', '기관_20일', '외국인_일간', '외국인_5일', '외국인_20일']:
            if col in result.columns:
                result[col] = (result[col] / 1000).round(1)

        # ticker_ks 컬럼 제거
        if 'ticker_ks' in result.columns:
            result = result.drop(columns=['ticker_ks'])

        # 스코어 계산
        print("스코어 계산 중...")
        result = self.calculate_scores(result)

        self._result_df = result

        print(f"최종 데이터: {len(result)}개 종목")

        # 등급별 분포 출력
        if '등급' in result.columns:
            grade_dist = result['등급'].value_counts().sort_index(ascending=False)
            print(f"등급 분포: {dict(grade_dist)}")

        return result

    # =========================================================================
    # Step 5: 스코어링 시스템
    # =========================================================================
    def _normalize_score(self, series: pd.Series, higher_is_better: bool = True) -> pd.Series:
        """0~100 점수로 정규화 (백분위 기반)"""
        if series.isna().all():
            return pd.Series([50] * len(series), index=series.index)

        pct_rank = series.rank(pct=True, na_option='keep')

        if higher_is_better:
            return (pct_rank * 100).round(1)
        else:
            return ((1 - pct_rank) * 100).round(1)

    def calculate_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        스코어링 시스템
        - 모멘텀 스코어 (30%): 20일수익률 + 10일선돌파 + RSI
        - 수급 스코어 (50%): 기관 + 외인 (일간/5일/20일)
        - 거래량 스코어 (20%): RVOL(20)
        """
        result = df.copy()

        # ===== 1. 모멘텀 스코어 (1일수익률 30% + 20일수익률 30% + 10일선돌파 20% + RSI 20%) =====
        momentum_parts = []
        momentum_weights = []

        if '1일수익률(%)' in result.columns:
            result['_m1'] = self._normalize_score(result['1일수익률(%)'], higher_is_better=True)
            momentum_parts.append('_m1')
            momentum_weights.append(0.3)

        if '20일수익률(%)' in result.columns:
            result['_m2'] = self._normalize_score(result['20일수익률(%)'], higher_is_better=True)
            momentum_parts.append('_m2')
            momentum_weights.append(0.3)

        if '10일선돌파' in result.columns:
            result['_m3'] = result['10일선돌파'].fillna(0) * 100
            momentum_parts.append('_m3')
            momentum_weights.append(0.2)

        if 'RSI(14)' in result.columns:
            result['_m4'] = self._normalize_score(result['RSI(14)'], higher_is_better=True)
            momentum_parts.append('_m4')
            momentum_weights.append(0.2)

        if momentum_parts:
            weighted = sum(result[c] * w for c, w in zip(momentum_parts, momentum_weights))
            total_w = sum(momentum_weights)
            result['모멘텀스코어'] = (weighted / total_w).round(1)
        else:
            result['모멘텀스코어'] = 50

        # ===== 2. 수급 스코어 =====
        supply_components = []

        # 기관
        if '기관_일간' in result.columns:
            result['_s1'] = self._normalize_score(result['기관_일간'], higher_is_better=True)
            supply_components.append(('_s1', self.config.SUPPLY_DAILY_WEIGHT))
        if '기관_5일' in result.columns:
            result['_s2'] = self._normalize_score(result['기관_5일'], higher_is_better=True)
            supply_components.append(('_s2', self.config.SUPPLY_5D_WEIGHT))
        if '기관_20일' in result.columns:
            result['_s3'] = self._normalize_score(result['기관_20일'], higher_is_better=True)
            supply_components.append(('_s3', self.config.SUPPLY_20D_WEIGHT))

        # 외국인
        if '외국인_일간' in result.columns:
            result['_s4'] = self._normalize_score(result['외국인_일간'], higher_is_better=True)
            supply_components.append(('_s4', self.config.SUPPLY_DAILY_WEIGHT))
        if '외국인_5일' in result.columns:
            result['_s5'] = self._normalize_score(result['외국인_5일'], higher_is_better=True)
            supply_components.append(('_s5', self.config.SUPPLY_5D_WEIGHT))
        if '외국인_20일' in result.columns:
            result['_s6'] = self._normalize_score(result['외국인_20일'], higher_is_better=True)
            supply_components.append(('_s6', self.config.SUPPLY_20D_WEIGHT))

        if supply_components:
            inst_cols = [c for c, _ in supply_components if c in ['_s1', '_s2', '_s3']]
            frgn_cols = [c for c, _ in supply_components if c in ['_s4', '_s5', '_s6']]

            inst_score = 50
            frgn_score = 50

            if inst_cols:
                inst_weights = [self.config.SUPPLY_DAILY_WEIGHT, self.config.SUPPLY_5D_WEIGHT, self.config.SUPPLY_20D_WEIGHT][:len(inst_cols)]
                inst_score = sum(result[c] * w for c, w in zip(inst_cols, inst_weights)) / sum(inst_weights)

            if frgn_cols:
                frgn_weights = [self.config.SUPPLY_DAILY_WEIGHT, self.config.SUPPLY_5D_WEIGHT, self.config.SUPPLY_20D_WEIGHT][:len(frgn_cols)]
                frgn_score = sum(result[c] * w for c, w in zip(frgn_cols, frgn_weights)) / sum(frgn_weights)

            result['수급스코어'] = ((inst_score + frgn_score) / 2).round(1)
        else:
            result['수급스코어'] = 50

        # ===== 3. 거래량 스코어 =====
        if 'RVOL(20)' in result.columns:
            result['거래량스코어'] = self._normalize_score(result['RVOL(20)'], higher_is_better=True)
        else:
            result['거래량스코어'] = 50

        # ===== 4. 종합 스코어 =====
        result['종합스코어'] = (
            result['모멘텀스코어'] * self.config.MOMENTUM_WEIGHT +
            result['수급스코어'] * self.config.SUPPLY_WEIGHT +
            result['거래량스코어'] * self.config.VOLUME_WEIGHT
        ).round(1)

        # 등급 부여
        grade = pd.cut(
            result['종합스코어'],
            bins=[0, 30, 50, 65, 80, 100],
            labels=['F', 'D', 'C', 'B', 'A'],
            include_lowest=True
        )
        # Categorical을 문자열로 변환 후 NaN 처리
        result['등급'] = grade.astype(str)
        result.loc[result['종합스코어'].isna(), '등급'] = '-'

        # 임시 컬럼 삭제
        temp_cols = [c for c in result.columns if c.startswith('_')]
        result = result.drop(columns=temp_cols)

        return result

    def _create_recommendation(self, df: pd.DataFrame) -> pd.DataFrame:
        """종합스코어 기반 추천종목 선정"""
        if '종합스코어' not in df.columns:
            return None

        top = df.nlargest(self.config.TOP_N_RESULTS, '종합스코어').copy()

        display_cols = [
            '코드', '코드명', '상장일', '현재가', '거래대금(억)',
            '종합스코어', '모멘텀스코어', '수급스코어', '거래량스코어',
            'RSI변화량', 'RSI(14)', '전일RSI(14)',
            'RVOL(20)', '기관_일간', '외국인_일간',
            '10일선돌파', '20일선돌파', '60일선돌파',
            '변동성(30D)', 'BB폭', '유통비율(%)', '공매도비율',
        ]
        display_cols = [c for c in display_cols if c in top.columns]

        return top[display_cols].reset_index(drop=True)

    # =========================================================================
    # Step 6: 결과 저장 및 출력
    # =========================================================================
    def _create_report_sheet(self, writer, df: pd.DataFrame, has_prev_data: bool = False):
        """POST IPO 보고서 시트 생성 (서식 + 파이차트)"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter

        wb = writer.book
        ws = wb.create_sheet('POST IPO 보고서', 0)  # 첫 번째 탭

        # ── 색상 정의 ──────────────────────────────────────────
        COLOR_NAVY   = '1F3864'
        COLOR_BLUE   = '2E75B6'
        COLOR_LIGHT  = 'D6E4F0'
        COLOR_WHITE  = 'FFFFFF'
        COLOR_GOLD   = 'C9A84C'
        COLOR_GREEN  = '70AD47'
        COLOR_ORANGE = 'ED7D31'
        COLOR_RED    = 'FF0000'
        COLOR_GRAY   = 'F2F2F2'

        def fill(color): return PatternFill('solid', fgColor=color)
        def font(bold=False, color='000000', size=11):
            return Font(bold=bold, color=color, size=size, name='맑은 고딕')
        def center(): return Alignment(horizontal='center', vertical='center')
        def thin_border():
            s = Side(style='thin', color='BFBFBF')
            return Border(left=s, right=s, top=s, bottom=s)

        ref_date_str = self._ref_date.strftime('%Y-%m-%d') if self._ref_date else datetime.now().strftime('%Y-%m-%d')
        total = len(df)

        # ── 등급 분포 계산 ──────────────────────────────────────
        grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
        if '등급' in df.columns:
            for g, cnt in df['등급'].value_counts().items():
                if g in grade_counts:
                    grade_counts[g] = int(cnt)

        # 모멘텀 구분: 강세(A+B), 중립(C), 약세(D+F)
        strong  = grade_counts['A'] + grade_counts['B']
        neutral = grade_counts['C']
        weak    = grade_counts['D'] + grade_counts['F']

        # 종합스코어 50점 이상 종목 수
        score_50_plus = int(
            (pd.to_numeric(df.get('종합스코어', pd.Series(dtype=float)), errors='coerce') >= 50).sum()
        )

        # ── 컬럼 너비 설정 ─────────────────────────────────────
        col_widths = {1: 3, 2: 18, 3: 14, 4: 10, 5: 10, 6: 10, 7: 10, 8: 12, 9: 12, 10: 10, 11: 14, 12: 3, 13: 16, 14: 10}
        for col, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # ── 타이틀 영역 ────────────────────────────────────────
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 38
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 12

        ws.merge_cells('B2:I2')
        title_cell = ws['B2']
        title_cell.value = 'POST IPO 모니터링 보고서'
        title_cell.font = Font(bold=True, color=COLOR_WHITE, size=18, name='맑은 고딕')
        title_cell.fill = fill(COLOR_NAVY)
        title_cell.alignment = center()

        ws.merge_cells('B3:I3')
        sub_cell = ws['B3']
        sub_cell.value = f'기준일: {ref_date_str}  |  분석 대상: {total}개 종목  |  2년 이내 신규 상장주'
        sub_cell.font = Font(color=COLOR_WHITE, size=10, name='맑은 고딕')
        sub_cell.fill = fill(COLOR_BLUE)
        sub_cell.alignment = center()

        # ── 핵심 지표 카드 (행 5~8) ────────────────────────────
        ws.row_dimensions[5].height = 10
        ws.row_dimensions[6].height = 28
        ws.row_dimensions[7].height = 22
        ws.row_dimensions[8].height = 12

        cards = [
            ('전체 종목', str(total), COLOR_BLUE),
            ('A등급', str(grade_counts['A']), COLOR_GOLD),
            ('B등급', str(grade_counts['B']), COLOR_GREEN),
            ('50점 이상\n종목수', str(score_50_plus), COLOR_NAVY),
        ]
        card_cols = [2, 4, 6, 8]
        for (label, value, color), col in zip(cards, card_cols):
            # 병합 먼저 실행
            ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
            ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)

            # 라벨 셀
            lc = ws.cell(row=6, column=col)
            lc.value = label
            lc.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
            lc.fill = fill(color)
            lc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 값 셀
            vc = ws.cell(row=7, column=col)
            vc.value = int(value)
            vc.font = Font(bold=True, color=color, size=20, name='맑은 고딕')
            vc.fill = fill(COLOR_GRAY)
            vc.alignment = center()

        # ── 섹션: 추천종목 TOP 10 (행 10~) ────────────────────
        ws.row_dimensions[9].height = 10
        ws.row_dimensions[10].height = 22

        ws.merge_cells('B10:O10')
        sec1 = ws['B10']
        sec1.value = '▶  추천종목  (종합스코어 50점 이상)'
        sec1.font = Font(bold=True, color=COLOR_WHITE, size=11, name='맑은 고딕')
        sec1.fill = fill(COLOR_BLUE)
        sec1.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        headers = ['순위', '종목명', '종합점수', '모멘텀', '수급', '거래량', 'RSI', '등급', 'RSI신호']
        header_cols = list(range(2, 11))  # B~J (columns 2~10)
        ws.row_dimensions[11].height = 18
        for h, c in zip(headers, header_cols):
            cell = ws.cell(row=11, column=c)
            cell.value = h
            cell.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
            cell.fill = fill(COLOR_NAVY)
            cell.alignment = center()
            cell.border = thin_border()

        # 추천종목 데이터: 종합스코어 50점 이상 전체 (TOP_N 제한 없이 df에서 직접 필터)
        grade_colors = {'A': 'FFD700', 'B': COLOR_GREEN, 'C': 'FFC000', 'D': 'ED7D31', 'F': 'FF6B6B'}

        if '종합스코어' in df.columns:
            scored = pd.to_numeric(df['종합스코어'], errors='coerce')
            filtered_rec = df[scored >= 50].sort_values('종합스코어', ascending=False).reset_index(drop=True)
        else:
            filtered_rec = pd.DataFrame()

        if not filtered_rec.empty:
            for idx, (_, row) in enumerate(filtered_rec.iterrows()):
                r = 12 + idx
                ws.row_dimensions[r].height = 17
                row_fill = fill(COLOR_WHITE) if idx % 2 == 0 else fill(COLOR_GRAY)
                grade_str = row.get('등급', '-') if pd.notna(row.get('등급', None)) else '-'

                # RSI 65 신규 돌파 신호: 전일 RSI < 65 이고 당일 RSI >= 65
                rsi_cur  = pd.to_numeric(row.get('RSI(14)', None), errors='coerce')
                rsi_prev_val = pd.to_numeric(row.get('전일RSI(14)', None), errors='coerce')
                rsi_signal = (
                    '★' if (pd.notna(rsi_cur) and pd.notna(rsi_prev_val)
                            and rsi_cur >= 65 and rsi_prev_val < 65)
                    else ''
                )

                data = [
                    idx + 1,
                    row.get('코드명', ''),
                    row.get('종합스코어', ''),
                    row.get('모멘텀스코어', ''),
                    row.get('수급스코어', ''),
                    row.get('거래량스코어', ''),
                    row.get('RSI(14)', ''),
                    grade_str,
                    rsi_signal,
                ]
                for val, c in zip(data, header_cols):
                    cell = ws.cell(row=r, column=c)
                    cell.value = val if not (isinstance(val, float) and pd.isna(val)) else '-'
                    cell.font = font(size=9)
                    cell.fill = row_fill
                    cell.alignment = center()
                    cell.border = thin_border()

                # 등급 셀 색상 (column 9)
                grade_cell = ws.cell(row=r, column=9)
                g_color = grade_colors.get(grade_str, 'BFBFBF')
                grade_cell.fill = fill(g_color)
                grade_cell.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')

                # RSI 신호 셀 색상 (column 10) — 신호 있을 때 강조
                if rsi_signal:
                    sig_cell = ws.cell(row=r, column=10)
                    sig_cell.fill = fill('FFF2CC')  # 연노랑
                    sig_cell.font = Font(bold=True, color='C9A84C', size=10, name='맑은 고딕')

        # ── K·L열: 등급변화 + 점수변화량 (전일 데이터 있을 때) ─────
        if has_prev_data:
            # K11 헤더
            k11 = ws.cell(row=11, column=11)
            k11.value = '등급변화'
            k11.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
            k11.fill = fill(COLOR_NAVY)
            k11.alignment = center()
            k11.border = thin_border()

            # L11 헤더
            l11 = ws.cell(row=11, column=12)
            l11.value = '점수변화량'
            l11.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
            l11.fill = fill(COLOR_NAVY)
            l11.alignment = center()
            l11.border = thin_border()

            # K12+, L12+ 수식 (추천종목 데이터 행)
            # INDEX/MATCH 사용 — 컬럼 수가 변해도 컬럼명으로 정확히 참조
            # 전체_전일자 시트: 1행=헤더, B열=코드명
            rec_count = len(filtered_rec) if not filtered_rec.empty else 0
            for idx in range(rec_count):
                r = 12 + idx
                row_fill = fill(COLOR_WHITE) if idx % 2 == 0 else fill(COLOR_GRAY)

                # 등급변화: INDEX/MATCH로 전일 등급 찾기
                k_cell = ws.cell(row=r, column=11)
                prev_grade = (
                    'INDEX(전체_전일자!1:1048576,'
                    'MATCH(C{r},전체_전일자!B:B,0),'
                    'MATCH("등급",전체_전일자!1:1,0))'
                ).format(r=r)
                k_cell.value = (
                    f'=IFERROR(IF(I{r}={prev_grade},'
                    f'"유지",'
                    f'{prev_grade}&" → "&I{r}),"")'
                )
                k_cell.font = font(size=9)
                k_cell.fill = row_fill
                k_cell.alignment = center()
                k_cell.border = thin_border()

                # 점수변화량: 당일 종합점수 - 전일 종합스코어
                l_cell = ws.cell(row=r, column=12)
                prev_score = (
                    'INDEX(전체_전일자!1:1048576,'
                    'MATCH(C{r},전체_전일자!B:B,0),'
                    'MATCH("종합스코어",전체_전일자!1:1,0))'
                ).format(r=r)
                l_cell.value = f'=IFERROR(D{r}-{prev_score},"")'
                l_cell.number_format = '+0.0;-0.0;0.0'
                l_cell.font = font(size=9)
                l_cell.fill = row_fill
                l_cell.alignment = center()
                l_cell.border = thin_border()

            # M·N·O열: 모멘텀/수급/거래량 변화량
            for col_idx, label in [(13, '모멘텀Δ'), (14, '수급Δ'), (15, '거래량Δ')]:
                hdr = ws.cell(row=11, column=col_idx)
                hdr.value = label
                hdr.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
                hdr.fill = fill(COLOR_NAVY)
                hdr.alignment = center()
                hdr.border = thin_border()

            # M12+, N12+, O12+ 수식 — INDEX/MATCH로 컬럼명 기반 참조
            score_cols = [
                (13, 'E', '모멘텀스코어'),
                (14, 'F', '수급스코어'),
                (15, 'G', '거래량스코어'),
            ]
            for idx in range(rec_count):
                r = 12 + idx
                row_fill_val = fill(COLOR_WHITE) if idx % 2 == 0 else fill(COLOR_GRAY)
                for col_idx, src_col, prev_col_name in score_cols:
                    prev_val = (
                        'INDEX(전체_전일자!1:1048576,'
                        'MATCH(C{r},전체_전일자!B:B,0),'
                        'MATCH("{col}",전체_전일자!1:1,0))'
                    ).format(r=r, col=prev_col_name)
                    cell = ws.cell(row=r, column=col_idx)
                    cell.value = f'=IFERROR({src_col}{r}-{prev_val},"")'
                    cell.number_format = '+0.0;-0.0;0.0'
                    cell.font = font(size=9)
                    cell.fill = row_fill_val
                    cell.alignment = center()
                    cell.border = thin_border()

            # 컬럼 너비
            ws.column_dimensions['K'].width = 14
            ws.column_dimensions['L'].width = 16
            ws.column_dimensions['M'].width = 11
            ws.column_dimensions['N'].width = 10
            ws.column_dimensions['O'].width = 11

        # ── J6~J7: RSI>65 라벨 + COUNTIF ──────────────────────────
        ws.merge_cells(start_row=6, start_column=10, end_row=6, end_column=10)
        j6 = ws.cell(row=6, column=10)
        j6.value = 'RSI>65'
        j6.font = Font(bold=True, color=COLOR_WHITE, size=9, name='맑은 고딕')
        j6.fill = fill(COLOR_NAVY)
        j6.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        j7 = ws.cell(row=7, column=10)
        # 컬럼 위치를 하드코딩(P열)하지 않고, 다른 수식들처럼 컬럼명으로 RSI(14) 열을 찾아 COUNTIF
        j7.value = (
            '=COUNTIF('
            'INDEX(전체!$A:$XFD,0,MATCH("RSI(14)",전체!$1:$1,0)),'
            '">=65")'
        )
        j7.font = Font(bold=True, color=COLOR_NAVY, size=20, name='맑은 고딕')
        j7.fill = fill(COLOR_GRAY)
        j7.alignment = center()

        # ── 파이차트 데이터: Q10 (column 17) 기준 ───────────────
        CHART_ROW = 10
        CHART_COL_LABEL = 17  # Q열
        CHART_COL_VAL   = 18  # R열

        ws.cell(row=CHART_ROW,     column=CHART_COL_LABEL).value = '구분'
        ws.cell(row=CHART_ROW,     column=CHART_COL_VAL).value   = '종목수'
        ws.cell(row=CHART_ROW + 1, column=CHART_COL_LABEL).value = '강세 (A+B등급)'
        ws.cell(row=CHART_ROW + 1, column=CHART_COL_VAL).value   = strong
        ws.cell(row=CHART_ROW + 2, column=CHART_COL_LABEL).value = '중립 (C등급)'
        ws.cell(row=CHART_ROW + 2, column=CHART_COL_VAL).value   = neutral
        ws.cell(row=CHART_ROW + 3, column=CHART_COL_LABEL).value = '약세 (D+F등급)'
        ws.cell(row=CHART_ROW + 3, column=CHART_COL_VAL).value   = weak

        ws.column_dimensions['Q'].width = 16
        ws.column_dimensions['R'].width = 10

        # 차트 생성
        pie = PieChart()
        pie.title = 'Post IPO 모멘텀 현황'
        pie.style = 10

        data_ref  = Reference(ws, min_col=CHART_COL_VAL,   min_row=CHART_ROW,     max_row=CHART_ROW + 3)
        label_ref = Reference(ws, min_col=CHART_COL_LABEL, min_row=CHART_ROW + 1, max_row=CHART_ROW + 3)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(label_ref)

        # 조각 색상 (강세=파랑, 중립=주황, 약세=빨강)
        slice_colors = [COLOR_BLUE, COLOR_ORANGE, 'C00000']
        for i, color in enumerate(slice_colors):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            pie.series[0].dPt.append(pt)

        from openpyxl.chart.label import DataLabelList
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = False

        pie.width  = 13
        pie.height = 10
        ws.add_chart(pie, 'Q14')

    def _set_column_width(self, writer, sheet_name: str):
        """Excel 시트의 모든 컬럼 폭을 설정"""
        from openpyxl.utils import get_column_letter

        excel_width = self.config.EXCEL_COLUMN_WIDTH / 7

        ws = writer.sheets[sheet_name]
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = excel_width

    def _load_rsi_history_from_excel(self) -> pd.DataFrame:
        """이전 엑셀 파일에서 RSI 히스토리 로드"""
        import glob

        pattern = str(self.config.OUTPUT_DIR / "ipo_monitoring_*.xlsx")
        files = glob.glob(pattern)

        if not files:
            return pd.DataFrame(columns=['날짜', '분석종목수', 'RSI65이상', '과매수비율(%)', '전일대비'])

        latest_file = max(files)

        try:
            df = pd.read_excel(latest_file, sheet_name='RSI추이')
            df['날짜'] = pd.to_datetime(df['날짜'])
            return df
        except (ValueError, KeyError):
            return pd.DataFrame(columns=['날짜', '분석종목수', 'RSI65이상', '과매수비율(%)', '전일대비'])

    def _calculate_rsi_history(self) -> pd.DataFrame:
        """RSI 65 이상 종목수 일별 히스토리 (Step 3에서 캐시한 RSI 시계열 재사용)"""
        cols = ['날짜', '분석종목수', 'RSI65이상', '과매수비율(%)', '전일대비']

        end_date = self._ref_date if self._ref_date else datetime.now()
        end_str = end_date.strftime('%Y-%m-%d')

        # 1. 기존 히스토리 로드
        existing = self._load_rsi_history_from_excel()
        is_initial = len(existing) < 5

        if is_initial:
            print(f"\nRSI 히스토리 초기화 ({self.config.RSI_CALC_DAYS}일 계산)...")
        else:
            print(f"\nRSI 히스토리 업데이트 (당일)...")

        # 2. 캐시된 RSI 시계열로 날짜×종목 행렬 구성 (네트워크 재요청 없음)
        if not self._rsi_cache:
            print("  RSI 캐시 없음")
            return existing if not existing.empty else pd.DataFrame(columns=cols)

        try:
            rsi_matrix = pd.DataFrame(self._rsi_cache).sort_index()
            rsi_matrix = rsi_matrix[rsi_matrix.index <= pd.to_datetime(end_date)]
            if rsi_matrix.empty:
                return existing if not existing.empty else pd.DataFrame(columns=cols)

            if is_initial:
                target_dates = list(rsi_matrix.index)[-self.config.RSI_CALC_DAYS:]
            else:
                target_dates = [rsi_matrix.index[-1]]

            calc_results = []
            for date in target_dates:
                row_data = rsi_matrix.loc[date]
                valid_count = int(row_data.notna().sum())
                overbought_count = int((row_data >= self.config.RSI_THRESHOLD).sum())
                ratio = (overbought_count / valid_count * 100) if valid_count > 0 else 0
                calc_results.append({
                    '날짜': pd.to_datetime(date),
                    '분석종목수': valid_count,
                    'RSI65이상': overbought_count,
                    '과매수비율(%)': round(ratio, 1),
                })

            calculated = pd.DataFrame(calc_results)

            if is_initial:
                print(f"  초기 {len(calculated)}일 계산 완료")
                history = calculated
            else:
                print(f"  당일 데이터 추가")
                existing = existing[existing['날짜'].dt.strftime('%Y-%m-%d') != end_str]
                history = pd.concat([existing, calculated], ignore_index=True)

            # 날짜순 정렬
            history = history.sort_values('날짜').reset_index(drop=True)

            # 90일 롤링 유지
            cutoff_date = end_date - pd.Timedelta(days=self.config.RSI_HISTORY_DAYS)
            history = history[history['날짜'] >= cutoff_date]

            # 전일대비 계산
            history['전일대비'] = history['RSI65이상'].diff().fillna(0).astype(int)

            print(f"  RSI 히스토리 총 {len(history)}일")
            return history

        except Exception as e:
            print(f"  RSI 히스토리 계산 오류: {e}")
            return existing if not existing.empty else pd.DataFrame(columns=cols)

    def _load_previous_day_data(self, prev_file: str = None) -> pd.DataFrame:
        """전일 결과 파일에서 '전체' 시트 로드 (파일명을 input으로 받음)"""
        try:
            if prev_file is None:
                prev_file = input("전일 결과 파일명을 입력하세요 (예: ipo_monitoring_20260326.xlsx): ").strip()

            if not prev_file:
                print("  전일 파일 미입력 → 전일 비교 생략")
                return None

            prev_path = self.config.OUTPUT_DIR / prev_file
            if not prev_path.exists():
                print(f"  파일 없음: {prev_path}")
                return None

            print(f"  전일 데이터 로드: {prev_file}")
            return pd.read_excel(prev_path, sheet_name='전체')
        except Exception as e:
            print(f"  전일 데이터 로드 실패: {e}")
            return None

    def save_results(self, filename: str = None, prev_file: str = None) -> Path:
        """결과를 Excel로 저장 (prev_file: 전일 결과 파일명, 미입력 시 input으로 받음)"""
        print("\n" + "=" * 60)
        print("[Step 6] 결과 저장")
        print("=" * 60)

        if self._result_df is None or self._result_df.empty:
            print("저장할 데이터가 없습니다.")
            return None

        df = self._result_df.copy()

        if filename is None:
            filename = self.config.get_output_filename()

        filepath = self.config.OUTPUT_DIR / filename

        sheet_names = []

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 추천종목
            recommend_df = self._create_recommendation(df)
            if recommend_df is not None and not recommend_df.empty:
                recommend_df.to_excel(writer, sheet_name='추천종목', index=False)
                sheet_names.append('추천종목')

            # A등급
            if '등급' in df.columns:
                grade_a = df[df['등급'] == 'A'].sort_values('종합스코어', ascending=False)
                if not grade_a.empty:
                    grade_a.to_excel(writer, sheet_name='A등급', index=False)
                    sheet_names.append('A등급')

            # 전체
            df.to_excel(writer, sheet_name='전체', index=False)
            sheet_names.append('전체')

            # 전체_전일자 시트 추가 (전일 결과 파일에서 로드)
            prev_df = self._load_previous_day_data(prev_file)
            has_prev_data = prev_df is not None and not prev_df.empty
            if has_prev_data:
                prev_df.to_excel(writer, sheet_name='전체_전일자', index=False)
                sheet_names.append('전체_전일자')

            # RSI 과매도/과매수
            if 'RSI(14)' in df.columns:
                oversold = df[df['RSI(14)'] < 30].sort_values('RSI(14)')
                if not oversold.empty:
                    oversold.to_excel(writer, sheet_name='RSI과매도', index=False)
                    sheet_names.append('RSI과매도')

                overbought = df[df['RSI(14)'] > 70].sort_values('RSI(14)', ascending=False)
                if not overbought.empty:
                    overbought.to_excel(writer, sheet_name='RSI과매수', index=False)
                    sheet_names.append('RSI과매수')

            # 외국인/기관 순매수 TOP
            if '외국인_일간' in df.columns:
                foreign_top = df.dropna(subset=['외국인_일간']).nlargest(20, '외국인_일간')
                if not foreign_top.empty:
                    foreign_top.to_excel(writer, sheet_name='외국인순매수TOP', index=False)
                    sheet_names.append('외국인순매수TOP')

            if '기관_일간' in df.columns:
                inst_top = df.dropna(subset=['기관_일간']).nlargest(20, '기관_일간')
                if not inst_top.empty:
                    inst_top.to_excel(writer, sheet_name='기관순매수TOP', index=False)
                    sheet_names.append('기관순매수TOP')

            # RVOL TOP
            if 'RVOL(20)' in df.columns:
                vol_top = df.dropna(subset=['RVOL(20)']).nlargest(20, 'RVOL(20)')
                if not vol_top.empty:
                    vol_top.to_excel(writer, sheet_name='RVOL_TOP', index=False)
                    sheet_names.append('RVOL_TOP')

            # RSI 히스토리 (RSI 65 이상 종목수 일별 추이, 30일)
            rsi_history = self._calculate_rsi_history()
            if not rsi_history.empty:
                rsi_display = rsi_history.copy()
                rsi_display['날짜'] = rsi_display['날짜'].dt.strftime('%Y-%m-%d')
                rsi_display = rsi_display.sort_values('날짜', ascending=False)
                rsi_display.to_excel(writer, sheet_name='RSI추이', index=False)
                sheet_names.append('RSI추이')

            # POST IPO 보고서 시트 (첫 번째 탭으로 추가)
            self._create_report_sheet(writer, df, has_prev_data=has_prev_data)

            # 컬럼 폭 적용
            for sheet in sheet_names:
                self._set_column_width(writer, sheet)

        ref_date_str = self._ref_date.strftime('%Y-%m-%d') if self._ref_date else '(미정)'
        print(f"저장 완료: {filepath}")
        print(f"  (데이터 기준일: {ref_date_str})")

        return filepath

    def print_summary(self, n: int = 10):
        """요약 출력"""
        if self._result_df is None or self._result_df.empty:
            print("데이터가 없습니다.")
            return

        df = self._result_df.copy()

        print("\n" + "=" * 80)
        print("                    IPO 모니터링 요약")
        print("=" * 80)
        ref_date_str = self._ref_date.strftime('%Y-%m-%d') if self._ref_date else '(미정)'
        print(f"데이터 기준일: {ref_date_str} (당일 종가)")
        print(f"실행일: {datetime.now().strftime('%Y-%m-%d')}")
        print(f"총 종목 수: {len(df)}개")

        if '등급' in df.columns:
            grade_dist = df['등급'].value_counts().sort_index(ascending=False)
            print(f"\n등급 분포: {dict(grade_dist)}")

        # 추천종목 TOP
        recommend = self._create_recommendation(df)
        if recommend is not None and not recommend.empty:
            print(f"\n[★ 추천종목 TOP {min(n, len(recommend))}] (종합스코어 기준)")
            print("-" * 80)
            print(f"  {'순위':>4} {'종목명':<12} {'종합':>6} {'모멘텀':>6} {'수급':>6} {'거래량':>6} {'RSI':>6} {'RSI변화':>7}")
            print("-" * 80)
            for i, (_, row) in enumerate(recommend.head(n).iterrows(), 1):
                print(f"  {i:>4} {row['코드명']:<12} "
                      f"{row['종합스코어']:>6.1f} {row.get('모멘텀스코어', 0):>6.1f} "
                      f"{row.get('수급스코어', 0):>6.1f} {row.get('거래량스코어', 0):>6.1f} "
                      f"{row.get('RSI(14)', 0):>6.1f} {row.get('RSI변화량', 0):>+7.1f}")

        # RSI 과매도
        if 'RSI(14)' in df.columns:
            oversold = df[df['RSI(14)'] < 30].sort_values('RSI(14)')
            if not oversold.empty:
                print(f"\n[RSI 과매도 (< 30)] {len(oversold)}개")
                print("-" * 70)
                for _, row in oversold.head(n).iterrows():
                    print(f"  {row['코드명']:<15} RSI: {row['RSI(14)']:>5.1f}  변동성: {row.get('변동성(30D)', 0):>5.1f}%")

        # 외국인 순매수 TOP
        if '외국인_일간' in df.columns:
            foreign_top = df.dropna(subset=['외국인_일간']).nlargest(n, '외국인_일간')
            if not foreign_top.empty:
                print(f"\n[외국인 순매수 TOP {n}] (단위: 백만원)")
                print("-" * 70)
                for _, row in foreign_top.iterrows():
                    print(f"  {row['코드명']:<15} 일간: {row['외국인_일간']:>8.1f}  5일: {row.get('외국인_5일', 0):>8.1f}  20일: {row.get('외국인_20일', 0):>8.1f}")

        # 기관 순매수 TOP
        if '기관_일간' in df.columns:
            inst_top = df.dropna(subset=['기관_일간']).nlargest(n, '기관_일간')
            if not inst_top.empty:
                print(f"\n[기관 순매수 TOP {n}] (단위: 백만원)")
                print("-" * 70)
                for _, row in inst_top.iterrows():
                    print(f"  {row['코드명']:<15} 일간: {row['기관_일간']:>8.1f}  5일: {row.get('기관_5일', 0):>8.1f}  20일: {row.get('기관_20일', 0):>8.1f}")

        # RVOL TOP
        if 'RVOL(20)' in df.columns:
            vol_top = df.dropna(subset=['RVOL(20)']).nlargest(n, 'RVOL(20)')
            if not vol_top.empty:
                print(f"\n[RVOL(20) TOP {n}]")
                print("-" * 70)
                for _, row in vol_top.iterrows():
                    print(f"  {row['코드명']:<15} RVOL: {row['RVOL(20)']:>5.2f}x  RSI: {row.get('RSI(14)', 0):>5.1f}")

        print("=" * 80)

    # =========================================================================
    # 전체 실행
    # =========================================================================
    def run(self, source: str = 'A', use_prev_day: bool = False) -> pd.DataFrame:
        """전체 모니터링 프로세스 실행
        source='A': 최초상장일.xlsx (기본)
        source='B': __post ipo univ.xlsx
        use_prev_day: True → 전일 종가 기준 (다음날 실행 시)
        """
        self.load_ipo_universe(source=source)
        self.load_supply_data()
        self.fetch_bloomberg_data(use_prev_day=use_prev_day)
        self.merge_data()
        self.print_summary()
        self.save_results()
        return self._result_df


# =============================================================================
# Main
# =============================================================================
if __name__ == "__main__":
    print("run.py를 통해 실행해주세요: python run.py")
